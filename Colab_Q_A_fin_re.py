"""
Automated Financial Report Generator for Tesla
Generates Excel reports with both Quarterly and Annual financial data
Uses yfinance to fetch data from Yahoo Finance
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import logging
import schedule
import time
import os
import sys # Import sys module

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class TeslaFinancialReportGenerator:
    """Generates financial reports for Tesla with quarterly and annual data"""

    def __init__(self, ticker="TSLA", output_file="tesla_financial_report.xlsx"):
        self.ticker = ticker
        self.output_file = output_file
        self.yf_ticker = yf.Ticker(ticker)

        # Fetch all financial data
        self.quarterly_balance_sheet = None
        self.annual_balance_sheet = None
        self.quarterly_income = None
        self.annual_income = None
        self.quarterly_cashflow = None
        self.annual_cashflow = None

    def fetch_all_data(self):
        """Fetch all financial data from yfinance"""
        logger.info(f"Fetching financial data for {self.ticker}...")

        try:
            # Fetch quarterly data
            self.quarterly_balance_sheet = self.yf_ticker.quarterly_balance_sheet
            self.quarterly_income = self.yf_ticker.quarterly_income_stmt
            self.quarterly_cashflow = self.yf_ticker.quarterly_cash_flow

            # Fetch annual data
            self.annual_balance_sheet = self.yf_ticker.balance_sheet
            self.annual_income = self.yf_ticker.income_stmt
            self.annual_cashflow = self.yf_ticker.cash_flow

            logger.info("Successfully fetched all financial data")
            return True

        except Exception as e:
            logger.error(f"Error fetching financial data: {e}")
            return False

    def create_excel_report(self):
        """Create Excel report with quarterly and annual data"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Define styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)

        # Title
        ws['A1'] = "FINANCIAL STATEMENTS"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:O1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Balance Sheet Section
        ws['A3'] = "Balance Sheet Data:"
        ws['A3'].font = header_font

        # Headers setup
        ws['B5'] = "In Thousands"
        ws['C5'] = 1000

        # Quarterly headers
        ws['E5'] = "Q"
        ws['G5'] = "Q"
        ws['I5'] = "Q"

        # Annual headers - changed 1/7/2025
        ws['K5'] = f"FY {datetime.now().year-1}"
        ws['M5'] = f"FY {datetime.now().year-2}"
        ws['O5'] = f"FY {datetime.now().year-3}"

        # Get dates for quarters
        if self.quarterly_balance_sheet is not None and len(self.quarterly_balance_sheet.columns) >= 3:
            q_dates = self.quarterly_balance_sheet.columns[:3]
            ws['E6'] = q_dates[0].strftime('%m/%d/%Y')
            ws['G6'] = q_dates[1].strftime('%m/%d/%Y')
            ws['I6'] = q_dates[2].strftime('%m/%d/%Y')

        # Get dates for annual
        if self.annual_balance_sheet is not None and len(self.annual_balance_sheet.columns) >= 3:
            a_dates = self.annual_balance_sheet.columns[:3]
            ws['K6'] = a_dates[0].strftime('%m/%d/%Y')
            ws['M6'] = a_dates[1].strftime('%m/%d/%Y')
            ws['O6'] = a_dates[2].strftime('%m/%d/%Y')

        # Assets header
        ws['B7'] = "Assets:"
        ws['B7'].font = subheader_font
        ws['F7'] = "Δ%"
        ws['H7'] = "Δ%"
        ws['J7'] = "Δ%"
        ws['L7'] = "Δ%"
        ws['N7'] = "Δ%"

        # Balance Sheet Items
        row = 8
        self._add_balance_sheet_items(ws, row)

        # Income Statement Section
        row = 31
        ws[f'A{row}'] = "Income Statement:"
        ws[f'A{row}'].font = header_font

        # Income Statement headers
        row = 32
        ws[f'K{row}'] = f"FY {datetime.now().year-1}"
        ws[f'M{row}'] = f"FY {datetime.now().year-2}"
        ws[f'O{row}'] = f"FY {datetime.now().year-3}"

        row = 35
        self._add_income_statement_items(ws, row)

        # Cash Flow Section
        row = 56
        ws[f'A{row}'] = "Cash Flows:"
        ws[f'A{row}'].font = header_font

        # Cash Flow headers
        row = 57
        ws[f'K{row}'] = f"FY {datetime.now().year-1}"
        ws[f'M{row}'] = f"FY {datetime.now().year-2}"
        ws[f'O{row}'] = f"FY {datetime.now().year-3}"

        row = 62
        self._add_cash_flow_items(ws, row)

        # Set column widths
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 4
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 8
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 8
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 8
        ws.column_dimensions['O'].width = 15

        # Save the workbook
        wb.save(self.output_file)
        logger.info(f"Report saved as {self.output_file}")

    def _add_balance_sheet_items(self, ws, start_row):
        """Add balance sheet items to the worksheet"""
        row = start_row

        # Define balance sheet items with their yfinance field names
        balance_sheet_items = [
            ("Cash and Equivalents", "Cash And Cash Equivalents"),
            ("Short-Term Investments", "Other Short Term Investments"),
            ("Accounts Receivable", "Accounts Receivable"),
            ("Inventories", "Inventory"),
            ("Current Assets", "Current Assets"),
            ("Total Assets", "Total Assets"),
            ("Working Capital", None),  # Calculated field
            
        ]

        for item_name, field_name in balance_sheet_items:
            ws[f'C{row}'] = item_name

            if field_name and field_name != "Working Capital":
                # Add quarterly data
                self._add_quarterly_data(ws, row, self.quarterly_balance_sheet, field_name, 'E', 'F', 'G', 'H', 'I', 'J')

                # Add annual data
                self._add_annual_data(ws, row, self.annual_balance_sheet, field_name, 'K', 'L', 'M', 'N', 'O')

            elif item_name == "Working Capital":
                # Calculate Working Capital = Current Assets - Current Liabilities
                self._calculate_working_capital(ws, row)

            row += 1

        # Add blank row
        row += 1

        # Liabilities section
        ws[f'B{row}'] = "Liabilities:"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        liability_items = [
            ("Short-Term Debt", "Short Term Debt"),
            ("Accounts Payable", "Accounts Payable"),
            ("Other Current Liabilities", "Other Current Liabilities"),
            ("Current Liabilities", "Current Liabilities"),
            ("Long-Term Debt", "Long Term Debt"),
            ("Total Liabilities", "Total Liabilities Net Minority Interest"),
            ("Net Worth (OE)", "Total Stockholder Equity"),
        ]

        for item_name, field_name in liability_items:
            ws[f'C{row}'] = item_name

            if field_name:
                # Add quarterly data
                self._add_quarterly_data(ws, row, self.quarterly_balance_sheet, field_name, 'E', 'F', 'G', 'H', 'I', 'J')

                # Add annual data
                self._add_annual_data(ws, row, self.annual_balance_sheet, field_name, 'K', 'L', 'M', 'N', 'O')

            row += 1

        # Add financial ratios
        row += 1

        # Current Ratio
        ws[f'A{row}'] = "Current Ratio"
        self._calculate_current_ratio(ws, row)
        row += 1

        # Quick Ratio
        ws[f'A{row}'] = "Quick Ratio"
        self._calculate_quick_ratio(ws, row)

        return row

    def _add_income_statement_items(self, ws, start_row):
        """Add income statement items to the worksheet"""
        row = start_row

        income_items = [
            ("Total Revenue", "Total Revenue"),
            ("Cost of Revenue", "Cost Of Revenue"),
            ("Gross Profit", "Gross Profit"),
            ("Operating Expenses", "Operating Expense"),
            ("Operating Income", "Operating Income"),
            ("EBITDA", "EBITDA"),
            ("EBIT", "EBIT"),
            ("Pretax Income", "Pretax Income"),
            ("Tax Provision", "Tax Provision"),
            ("Net Income", "Net Income"),
            ("Basic EPS", "Basic EPS"),
            ("Diluted EPS", "Diluted EPS"),
            ("Basic Average Shares", "Basic Average Shares"),
            ("Diluted Average Shares", "Diluted Average Shares"),
            ("Total Operating Expenses", "Total Expenses"),
            ("Interest Expense", "Interest Expense"),
            ("Gross Margin", None),  # Calculated
        ]

        for item_name, field_name in income_items:
            ws[f'C{row}'] = item_name

            if field_name:
                # Add quarterly data
                self._add_quarterly_data(ws, row, self.quarterly_income, field_name, 'E', 'F', 'G', 'H', 'I', 'J')

                # Add annual data
                self._add_annual_data(ws, row, self.annual_income, field_name, 'K', 'L', 'M', 'N', 'O')

            elif item_name == "Gross Margin":
                # Calculate Gross Margin = Gross Profit / Revenue
                self._calculate_gross_margin(ws, row)

            row += 1

        return row

    def _add_cash_flow_items(self, ws, start_row):
        """Add cash flow items to the worksheet"""
        row = start_row

        # Operating Activities
        ws[f'B{row}'] = "Cash Flows-Operating Activities:"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        operating_items = [
            ("Net Income", "Net Income"),
            ("Depreciation & Amortization", "Depreciation And Amortization"),
            ("Stock Based Compensation", "Stock Based Compensation"),
            ("Change in Working Capital", "Change In Working Capital"),
            ("Other Operating Activities", "Other Non Cash Items"),
        ]

        for item_name, field_name in operating_items:
            ws[f'C{row}'] = item_name

            if field_name:
                self._add_quarterly_data(ws, row, self.quarterly_cashflow, field_name, 'E', 'F', 'G', 'H', 'I', 'J')
                self._add_annual_data(ws, row, self.annual_cashflow, field_name, 'K', 'L', 'M', 'N', 'O')

            row += 1

        # Net Cash Flow-Operating
        ws[f'B{row}'] = "Net Cash Flow-Operating"
        self._add_quarterly_data(ws, row, self.quarterly_cashflow, "Operating Cash Flow", 'E', 'F', 'G', 'H', 'I', 'J')
        self._add_annual_data(ws, row, self.annual_cashflow, "Operating Cash Flow", 'K', 'L', 'M', 'N', 'O')
        row += 2

        # Investing Activities
        ws[f'B{row}'] = "Cash Flows-Investing Activities:"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        investing_items = [
            ("Capital Expenditures", "Capital Expenditure"),
            ("Acquisitions", "Net Business Purchase And Sale"),
            ("Investments", "Net Investment Purchase And Sale"),
            ("Other Investing Activities", "Net Other Investing Changes"),
        ]

        for item_name, field_name in investing_items:
            ws[f'C{row}'] = item_name

            if field_name:
                self._add_quarterly_data(ws, row, self.quarterly_cashflow, field_name, 'E', 'F', 'G', 'H', 'I', 'J')
                self._add_annual_data(ws, row, self.annual_cashflow, field_name, 'K', 'L', 'M', 'N', 'O')

            row += 1

        # Net Cash Flows-Investing
        ws[f'B{row}'] = "Net Cash Flows-Investing"
        self._add_quarterly_data(ws, row, self.quarterly_cashflow, "Investing Cash Flow", 'E', 'F', 'G', 'H', 'I', 'J')
        self._add_annual_data(ws, row, self.annual_cashflow, "Investing Cash Flow", 'K', 'L', 'M', 'N', 'O')
        row += 2

        # Financing Activities
        ws[f'B{row}'] = "Cash Flows-Financing Activities:"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        financing_items = [
            ("Debt Issuance/Retirement", "Net Issuance Payments Of Debt"),
            ("Stock Issuance/Buyback", "Net Common Stock Issuance"),
            ("Dividends Paid", "Cash Dividends Paid"),
            ("Other Financing Activities", "Net Other Financing Charges"),
        ]

        for item_name, field_name in financing_items:
            ws[f'C{row}'] = item_name

            if field_name:
                self._add_quarterly_data(ws, row, self.quarterly_cashflow, field_name, 'E', 'F', 'G', 'H', 'I', 'J')
                self._add_annual_data(ws, row, self.annual_cashflow, field_name, 'K', 'L', 'M', 'N', 'O')

            row += 1

        # Net Cash Flows-Financing
        ws[f'B{row}'] = "Net Cash Flows-Financing"
        self._add_quarterly_data(ws, row, self.quarterly_cashflow, "Financing Cash Flow", 'E', 'F', 'G', 'H', 'I', 'J')
        self._add_annual_data(ws, row, self.annual_cashflow, "Financing Cash Flow", 'K', 'L', 'M', 'N', 'O')
        row += 2

        # Net Cash Flow
        ws[f'B{row}'] = "Net Cash Flow"
        self._calculate_net_cash_flow(ws, row)

        return row

    def _add_quarterly_data(self, ws, row, dataframe, field_name, col1, col2, col3, col4, col5, col6):
        """Add quarterly data with percentage changes"""
        if dataframe is None or field_name not in dataframe.index:
            return

        try:
            values = dataframe.loc[field_name].values[:3]

            # Add values
            ws[f'{col1}{row}'] = self._format_currency(values[0])
            ws[f'{col3}{row}'] = self._format_currency(values[1])
            ws[f'{col5}{row}'] = self._format_currency(values[2])

            # Calculate and add percentage changes
            if len(values) >= 2:
                pct_change1 = self._calculate_pct_change(values[0], values[1])
                ws[f'{col2}{row}'] = pct_change1
                ws[f'{col2}{row}'].number_format = '0.00%' if abs(pct_change1) < 10 else '0.0'

            if len(values) >= 3:
                pct_change2 = self._calculate_pct_change(values[1], values[2])
                ws[f'{col4}{row}'] = pct_change2
                ws[f'{col4}{row}'].number_format = '0.00%' if abs(pct_change2) < 10 else '0.0'

        except Exception as e:
            logger.warning(f"Error adding quarterly data for {field_name}: {e}")

    def _add_annual_data(self, ws, row, dataframe, field_name, col1, col2, col3, col4, col5):
        """Add annual data with percentage changes"""
        if dataframe is None or field_name not in dataframe.index:
            return

        try:
            values = dataframe.loc[field_name].values[:3]

            # Add values
            ws[f'{col1}{row}'] = self._format_currency(values[0])
            ws[f'{col3}{row}'] = self._format_currency(values[1])
            ws[f'{col5}{row}'] = self._format_currency(values[2])

            # Calculate and add percentage changes
            if len(values) >= 2:
                pct_change1 = self._calculate_pct_change(values[0], values[1])
                ws[f'{col2}{row}'] = pct_change1
                ws[f'{col2}{row}'].number_format = '0.00%' if abs(pct_change1) < 10 else '0.0'

            if len(values) >= 3:
                pct_change2 = self._calculate_pct_change(values[1], values[2])
                ws[f'{col4}{row}'] = pct_change2
                ws[f'{col4}{row}'].number_format = '0.00%' if abs(pct_change2) < 10 else '0.0'

        except Exception as e:
            logger.warning(f"Error adding annual data for {field_name}: {e}")

    def _calculate_working_capital(self, ws, row):
        """Calculate Working Capital = Current Assets - Current Liabilities"""
        # This would need to reference the cells with current assets and liabilities
        # For now, we'll use the data directly
        if self.quarterly_balance_sheet is not None:
            try:
                current_assets = self.quarterly_balance_sheet.loc['Total Current Assets'].values[:3]
                current_liabilities = self.quarterly_balance_sheet.loc['Total Current Liabilities'].values[:3]

                for i, (ca, cl) in enumerate(zip(current_assets, current_liabilities)):
                    working_capital = ca - cl
                    col = ['E', 'G', 'I'][i]
                    ws[f'{col}{row}'] = self._format_currency(working_capital)

                    # Add percentage changes
                    if i > 0:
                        prev_wc = current_assets[i-1] - current_liabilities[i-1]
                        pct_change = self._calculate_pct_change(working_capital, prev_wc)
                        pct_col = ['F', 'H'][i-1]
                        ws[f'{pct_col}{row}'] = pct_change
                        ws[f'{pct_col}{row}'].number_format = '0.00%' if abs(pct_change) < 10 else '0.0'

            except Exception as e:
                logger.warning(f"Error calculating working capital: {e}")

        # Similar calculation for annual data
        if self.annual_balance_sheet is not None:
            try:
                current_assets = self.annual_balance_sheet.loc['Total Current Assets'].values[:3]
                current_liabilities = self.annual_balance_sheet.loc['Total Current Liabilities'].values[:3]

                for i, (ca, cl) in enumerate(zip(current_assets, current_liabilities)):
                    working_capital = ca - cl
                    col = ['K', 'M', 'O'][i]
                    ws[f'{col}{row}'] = self._format_currency(working_capital)

                    # Add percentage changes
                    if i > 0:
                        prev_wc = current_assets[i-1] - current_liabilities[i-1]
                        pct_change = self._calculate_pct_change(working_capital, prev_wc)
                        pct_col = ['L', 'N'][i-1]
                        ws[f'{pct_col}{row}'] = pct_change
                        ws[f'{pct_col}{row}'].number_format = '0.00%' if abs(pct_change) < 10 else '0.0'

            except Exception as e:
                logger.warning(f"Error calculating annual working capital: {e}")

    def _calculate_current_ratio(self, ws, row):
        """Calculate Current Ratio = Current Assets / Current Liabilities"""
        if self.quarterly_balance_sheet is not None:
            try:
                current_assets = self.quarterly_balance_sheet.loc['Total Current Assets'].values[:3]
                current_liabilities = self.quarterly_balance_sheet.loc['Total Current Liabilities'].values[:3]

                for i, (ca, cl) in enumerate(zip(current_assets, current_liabilities)):
                    if cl != 0:
                        ratio = ca / cl
                        col = ['E', 'G', 'I'][i]
                        ws[f'{col}{row}'] = round(ratio, 2)

            except Exception as e:
                logger.warning(f"Error calculating current ratio: {e}")

        # Annual data
        if self.annual_balance_sheet is not None:
            try:
                current_assets = self.annual_balance_sheet.loc['Total Current Assets'].values[:3]
                current_liabilities = self.annual_balance_sheet.loc['Total Current Liabilities'].values[:3]

                for i, (ca, cl) in enumerate(zip(current_assets, current_liabilities)):
                    if cl != 0:
                        ratio = ca / cl
                        col = ['K', 'M', 'O'][i]
                        ws[f'{col}{row}'] = round(ratio, 2)

            except Exception as e:
                logger.warning(f"Error calculating annual current ratio: {e}")

    def _calculate_quick_ratio(self, ws, row):
        """Calculate Quick Ratio = (Current Assets - Inventory) / Current Liabilities"""
        if self.quarterly_balance_sheet is not None:
            try:
                current_assets = self.quarterly_balance_sheet.loc['Total Current Assets'].values[:3]
                inventory = self.quarterly_balance_sheet.loc['Inventory'].values[:3] if 'Inventory' in self.quarterly_balance_sheet.index else [0, 0, 0]
                current_liabilities = self.quarterly_balance_sheet.loc['Total Current Liabilities'].values[:3]

                for i, (ca, inv, cl) in enumerate(zip(current_assets, inventory, current_liabilities)):
                    if cl != 0:
                        ratio = (ca - inv) / cl
                        col = ['E', 'G', 'I'][i]
                        ws[f'{col}{row}'] = round(ratio, 2)

            except Exception as e:
                logger.warning(f"Error calculating quick ratio: {e}")

        # Annual data
        if self.annual_balance_sheet is not None:
            try:
                current_assets = self.annual_balance_sheet.loc['Total Current Assets'].values[:3]
                inventory = self.annual_balance_sheet.loc['Inventory'].values[:3] if 'Inventory' in self.annual_balance_sheet.index else [0, 0, 0]
                current_liabilities = self.annual_balance_sheet.loc['Total Current Liabilities'].values[:3]

                for i, (ca, inv, cl) in enumerate(zip(current_assets, inventory, current_liabilities)):
                    if cl != 0:
                        ratio = (ca - inv) / cl
                        col = ['K', 'M', 'O'][i]
                        ws[f'{col}{row}'] = round(ratio, 2)

            except Exception as e:
                logger.warning(f"Error calculating annual quick ratio: {e}")

    def _calculate_gross_margin(self, ws, row):
        """Calculate Gross Margin = Gross Profit / Revenue"""
        if self.quarterly_income is not None:
            try:
                revenue = self.quarterly_income.loc['Total Revenue'].values[:3]
                gross_profit = self.quarterly_income.loc['Gross Profit'].values[:3]

                for i, (rev, gp) in enumerate(zip(revenue, gross_profit)):
                    if rev != 0:
                        margin = gp / rev
                        col = ['E', 'G', 'I'][i]
                        ws[f'{col}{row}'] = margin
                        ws[f'{col}{row}'].number_format = '0.0%'

            except Exception as e:
                logger.warning(f"Error calculating gross margin: {e}")

        # Annual data
        if self.annual_income is not None:
            try:
                revenue = self.annual_income.loc['Total Revenue'].values[:3]
                gross_profit = self.annual_income.loc['Gross Profit'].values[:3]

                for i, (rev, gp) in enumerate(zip(revenue, gross_profit)):
                    if rev != 0:
                        margin = gp / rev
                        col = ['K', 'M', 'O'][i]
                        ws[f'{col}{row}'] = margin
                        ws[f'{col}{row}'].number_format = '0.0%'

            except Exception as e:
                logger.warning(f"Error calculating annual gross margin: {e}")

    def _calculate_net_cash_flow(self, ws, row):
        """Calculate Net Cash Flow = Operating + Investing + Financing Cash Flow"""
        if self.quarterly_cashflow is not None:
            try:
                operating = self.quarterly_cashflow.loc['Operating Cash Flow'].values[:3] if 'Operating Cash Flow' in self.quarterly_cashflow.index else [0, 0, 0]
                investing = self.quarterly_cashflow.loc['Investing Cash Flow'].values[:3] if 'Investing Cash Flow' in self.quarterly_cashflow.index else [0, 0, 0]
                financing = self.quarterly_cashflow.loc['Financing Cash Flow'].values[:3] if 'Financing Cash Flow' in self.quarterly_cashflow.index else [0, 0, 0]

                for i, (op, inv, fin) in enumerate(zip(operating, investing, financing)):
                    net_cash = op + inv + fin
                    col = ['E', 'G', 'I'][i]
                    ws[f'{col}{row}'] = self._format_currency(net_cash)

            except Exception as e:
                logger.warning(f"Error calculating net cash flow: {e}")

        # Annual data
        if self.annual_cashflow is not None:
            try:
                operating = self.annual_cashflow.loc['Operating Cash Flow'].values[:3] if 'Operating Cash Flow' in self.annual_cashflow.index else [0, 0, 0]
                investing = self.annual_cashflow.loc['Investing Cash Flow'].values[:3] if 'Investing Cash Flow' in self.annual_cashflow.index else [0, 0, 0]
                financing = self.annual_cashflow.loc['Financing Cash Flow'].values[:3] if 'Financing Cash Flow' in self.annual_cashflow.index else [0, 0, 0]

                for i, (op, inv, fin) in enumerate(zip(operating, investing, financing)):
                    net_cash = op + inv + fin
                    col = ['K', 'M', 'O'][i]
                    ws[f'{col}{row}'] = self._format_currency(net_cash)

            except Exception as e:
                logger.warning(f"Error calculating annual net cash flow: {e}")

    def _format_currency(self, value):
        """Format value as currency string"""
        if pd.isna(value) or value == 0:
            return "$0"

        # Values are already in thousands
        if value < 0:
            return f"$({abs(value):,.0f})"
        else:
            return f"${value:,.0f}"

    def _calculate_pct_change(self, current, previous):
        """Calculate percentage change"""
        if previous == 0:
            return 0
        return (current - previous) / abs(previous)

    def generate_report(self):
        """Main method to generate the complete report"""
        logger.info(f"Starting report generation for {self.ticker}...")

        # Fetch all data
        if not self.fetch_all_data():
            logger.error("Failed to fetch financial data")
            return False

        # Create Excel report
        self.create_excel_report()

        logger.info("Report generation completed successfully!")
        return True


class FinancialReportAutomation:
    """Automation wrapper for scheduling and running reports"""

    def __init__(self, ticker="TSLA", output_dir="./reports"):
        self.ticker = ticker
        self.output_dir = output_dir

        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)

    def run_report(self):
        """Run a single report generation"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(self.output_dir, f"{self.ticker}_financial_report_{timestamp}.xlsx")

        generator = TeslaFinancialReportGenerator(self.ticker, output_file)
        success = generator.generate_report()

        if success:
            logger.info(f"Report saved to: {output_file}")
            return output_file
        else:
            logger.error("Report generation failed")
            return None

    def schedule_daily_report(self, time_str="09:00"):
        """Schedule daily report generation"""
        logger.info(f"Scheduling daily report generation at {time_str}")

        schedule.every().day.at(time_str).do(self.run_report)

        while True:
            schedule.run_pending()
            time.sleep(60)  # Check every minute

    def schedule_weekly_report(self, day="monday", time_str="09:00"):
        """Schedule weekly report generation"""
        logger.info(f"Scheduling weekly report generation on {day} at {time_str}")

        getattr(schedule.every(), day).at(time_str).do(self.run_report)

        while True:
            schedule.run_pending()
            time.sleep(60)


def main():
    """Main function to run the automation"""
    import argparse

    parser = argparse.ArgumentParser(description='Tesla Financial Report Generator')
    parser.add_argument('--ticker', default='TSLA', help='Stock ticker symbol')
    parser.add_argument('--output', default='./reports', help='Output directory for reports')
    parser.add_argument('--schedule', choices=['none', 'daily', 'weekly'], default='none',
                        help='Schedule type for automatic generation')
    parser.add_argument('--time', default='09:00', help='Time for scheduled reports (HH:MM)')
    parser.add_argument('--day', default='monday', help='Day for weekly reports')

    # Parse known arguments and ignore the rest
    args, unknown = parser.parse_known_args()

    # Create automation instance
    automation = FinancialReportAutomation(args.ticker, args.output)

    if args.schedule == 'none':
        # Run once
        automation.run_report()
    elif args.schedule == 'daily':
        # Schedule daily
        automation.schedule_daily_report(args.time)
    elif args.schedule == 'weekly':
        # Schedule weekly
        automation.schedule_weekly_report(args.day, args.time)


if __name__ == "__main__":
    main()