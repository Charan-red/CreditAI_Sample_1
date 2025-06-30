"""
Automated Financial Report Generator for Tesla
This script automates the process of scraping financial data and generating Excel reports
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import yfinance as yf
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import schedule
import time
import json
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class FinancialDataScraper:
    """Handles scraping financial data from various sources"""
    
    def __init__(self, ticker="TSLA"):
        self.ticker = ticker
        self.yf_ticker = yf.Ticker(ticker)
        
    def get_balance_sheet_data(self):
        """Fetch balance sheet data"""
        try:
            # Get quarterly balance sheet
            balance_sheet = self.yf_ticker.quarterly_balance_sheet
            
            # Extract key items
            data = {
                'dates': balance_sheet.columns.strftime('%m/%d/%Y').tolist()[:3],
                'cash_equivalents': balance_sheet.loc['Cash And Cash Equivalents'].tolist()[:3] if 'Cash And Cash Equivalents' in balance_sheet.index else [0,0,0],
                'short_term_investments': balance_sheet.loc['Other Short Term Investments'].tolist()[:3] if 'Other Short Term Investments' in balance_sheet.index else [0,0,0],
                'accounts_receivable': balance_sheet.loc['Accounts Receivable'].tolist()[:3] if 'Accounts Receivable' in balance_sheet.index else [0,0,0],
                'inventory': balance_sheet.loc['Inventory'].tolist()[:3] if 'Inventory' in balance_sheet.index else [0,0,0],
                'total_current_assets': balance_sheet.loc['Total Current Assets'].tolist()[:3] if 'Total Current Assets' in balance_sheet.index else [0,0,0],
                'ppe': balance_sheet.loc['Property Plant Equipment Net'].tolist()[:3] if 'Property Plant Equipment Net' in balance_sheet.index else [0,0,0],
                'total_assets': balance_sheet.loc['Total Assets'].tolist()[:3] if 'Total Assets' in balance_sheet.index else [0,0,0],
                'accounts_payable': balance_sheet.loc['Accounts Payable'].tolist()[:3] if 'Accounts Payable' in balance_sheet.index else [0,0,0],
                'short_term_debt': balance_sheet.loc['Short Term Debt'].tolist()[:3] if 'Short Term Debt' in balance_sheet.index else [0,0,0],
                'total_current_liabilities': balance_sheet.loc['Total Current Liabilities'].tolist()[:3] if 'Total Current Liabilities' in balance_sheet.index else [0,0,0],
                'long_term_debt': balance_sheet.loc['Long Term Debt'].tolist()[:3] if 'Long Term Debt' in balance_sheet.index else [0,0,0],
                'total_liabilities': balance_sheet.loc['Total Liabilities'].tolist()[:3] if 'Total Liabilities' in balance_sheet.index else [0,0,0],
                'total_equity': balance_sheet.loc['Total Stockholder Equity'].tolist()[:3] if 'Total Stockholder Equity' in balance_sheet.index else [0,0,0],
            }
            
            return data
            
        except Exception as e:
            logger.error(f"Error fetching balance sheet: {e}")
            return None
    
    def get_income_statement_data(self):
        """Fetch income statement data"""
        try:
            # Get quarterly income statement
            income_stmt = self.yf_ticker.quarterly_income_stmt
            
            data = {
                'revenue': income_stmt.loc['Total Revenue'].tolist()[:3] if 'Total Revenue' in income_stmt.index else [0,0,0],
                'cost_of_revenue': income_stmt.loc['Cost Of Revenue'].tolist()[:3] if 'Cost Of Revenue' in income_stmt.index else [0,0,0],
                'gross_profit': income_stmt.loc['Gross Profit'].tolist()[:3] if 'Gross Profit' in income_stmt.index else [0,0,0],
                'operating_expenses': income_stmt.loc['Operating Expense'].tolist()[:3] if 'Operating Expense' in income_stmt.index else [0,0,0],
                'operating_income': income_stmt.loc['Operating Income'].tolist()[:3] if 'Operating Income' in income_stmt.index else [0,0,0],
                'ebit': income_stmt.loc['EBIT'].tolist()[:3] if 'EBIT' in income_stmt.index else [0,0,0],
                'interest_expense': income_stmt.loc['Interest Expense'].tolist()[:3] if 'Interest Expense' in income_stmt.index else [0,0,0],
                'pretax_income': income_stmt.loc['Pretax Income'].tolist()[:3] if 'Pretax Income' in income_stmt.index else [0,0,0],
                'tax_expense': income_stmt.loc['Tax Provision'].tolist()[:3] if 'Tax Provision' in income_stmt.index else [0,0,0],
                'net_income': income_stmt.loc['Net Income'].tolist()[:3] if 'Net Income' in income_stmt.index else [0,0,0],
            }
            
            return data
            
        except Exception as e:
            logger.error(f"Error fetching income statement: {e}")
            return None
    
    def get_cash_flow_data(self):
        """Fetch cash flow data"""
        try:
            # Get quarterly cash flow
            cash_flow = self.yf_ticker.quarterly_cash_flow
            
            data = {
                # Operating activities
                'operating_cash_flow': cash_flow.loc['Operating Cash Flow'].tolist()[:3] if 'Operating Cash Flow' in cash_flow.index else [0,0,0],
                'depreciation': cash_flow.loc['Depreciation'].tolist()[:3] if 'Depreciation' in cash_flow.index else [0,0,0],
                
                # Investing activities
                'capex': cash_flow.loc['Capital Expenditure'].tolist()[:3] if 'Capital Expenditure' in cash_flow.index else [0,0,0],
                'investing_cash_flow': cash_flow.loc['Investing Cash Flow'].tolist()[:3] if 'Investing Cash Flow' in cash_flow.index else [0,0,0],
                
                # Financing activities
                'debt_repayment': cash_flow.loc['Repayment Of Debt'].tolist()[:3] if 'Repayment Of Debt' in cash_flow.index else [0,0,0],
                'dividends_paid': cash_flow.loc['Cash Dividends Paid'].tolist()[:3] if 'Cash Dividends Paid' in cash_flow.index else [0,0,0],
                'financing_cash_flow': cash_flow.loc['Financing Cash Flow'].tolist()[:3] if 'Financing Cash Flow' in cash_flow.index else [0,0,0],
                
                # Total
                'free_cash_flow': cash_flow.loc['Free Cash Flow'].tolist()[:3] if 'Free Cash Flow' in cash_flow.index else [0,0,0],
            }
            
            return data
            
        except Exception as e:
            logger.error(f"Error fetching cash flow: {e}")
            return None

class ExcelReportGenerator:
    """Handles Excel report generation with formatting"""
    
    def __init__(self, filename="tesla_financial_report.xlsx"):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Sheet1"
        
        # Define styles
        self.header_font = Font(bold=True, size=14)
        self.subheader_font = Font(bold=True, size=12)
        self.normal_font = Font(size=11)
        self.number_format = '#,##0'
        self.percent_format = '0.00%'
        
    def calculate_percentage_change(self, current, previous):
        """Calculate percentage change between two values"""
        if previous == 0:
            return 0
        return (current - previous) / abs(previous)
    
    def format_currency(self, value):
        """Format value as currency in thousands"""
        return f"${value:,.0f}"
    
    def create_report(self, balance_sheet_data, income_stmt_data, cash_flow_data):
        """Create the complete financial report"""
        
        # Title
        self.ws['B1'] = "FINANCIAL STATEMENTS"
        self.ws['B1'].font = Font(bold=True, size=16)
        
        # Balance Sheet Section
        self.ws['B3'] = "Balance Sheet Data:"
        self.ws['B3'].font = self.header_font
        
        # Set up headers
        self.ws['C5'] = "In Thousands"
        self.ws['D5'] = 1000
        
        # Quarter headers
        dates = balance_sheet_data['dates']
        self.ws['F5'] = "Q"
        self.ws['H5'] = "Q"
        self.ws['J5'] = "Q"
        self.ws['F6'] = dates[0]
        self.ws['H6'] = dates[1]
        self.ws['J6'] = dates[2]
        
        # Assets section
        row = 7
        self.ws[f'C{row}'] = "Assets:"
        self.ws[f'C{row}'].font = self.subheader_font
        self.ws[f'G{row}'] = "Δ%"
        self.ws[f'I{row}'] = "Δ%"
        
        # Asset items
        asset_items = [
            ("Cash and Equivalents", balance_sheet_data['cash_equivalents']),
            ("Short-Term Investments", balance_sheet_data['short_term_investments']),
            ("Accounts Receivable", balance_sheet_data['accounts_receivable']),
            ("Inventory", balance_sheet_data['inventory']),
            ("Total Current Assets", balance_sheet_data['total_current_assets']),
            ("Property Plant and Equipment (PP&E)", balance_sheet_data['ppe']),
            ("Total Assets", balance_sheet_data['total_assets']),
        ]
        
        row = 8
        for item_name, values in asset_items:
            self.ws[f'D{row}'] = item_name
            self.ws[f'F{row}'] = self.format_currency(values[0])
            self.ws[f'H{row}'] = self.format_currency(values[1])
            self.ws[f'J{row}'] = self.format_currency(values[2])
            
            # Calculate percentage changes
            if row > 8:  # Skip percentage for first row
                self.ws[f'G{row}'] = self.calculate_percentage_change(values[0], values[1])
                self.ws[f'I{row}'] = self.calculate_percentage_change(values[1], values[2])
                self.ws[f'G{row}'].number_format = self.percent_format
                self.ws[f'I{row}'].number_format = self.percent_format
            
            row += 1
        
        # Liabilities section
        row += 1
        self.ws[f'C{row}'] = "Liabilities:"
        self.ws[f'C{row}'].font = self.subheader_font
        row += 1
        
        liability_items = [
            ("Accounts Payable", balance_sheet_data['accounts_payable']),
            ("Short-Term Debt", balance_sheet_data['short_term_debt']),
            ("Total Current Liabilities", balance_sheet_data['total_current_liabilities']),
            ("Long-Term Debt", balance_sheet_data['long_term_debt']),
            ("Total Liabilities", balance_sheet_data['total_liabilities']),
        ]
        
        for item_name, values in liability_items:
            self.ws[f'D{row}'] = item_name
            self.ws[f'F{row}'] = self.format_currency(values[0])
            self.ws[f'H{row}'] = self.format_currency(values[1])
            self.ws[f'J{row}'] = self.format_currency(values[2])
            
            self.ws[f'G{row}'] = self.calculate_percentage_change(values[0], values[1])
            self.ws[f'I{row}'] = self.calculate_percentage_change(values[1], values[2])
            self.ws[f'G{row}'].number_format = self.percent_format
            self.ws[f'I{row}'].number_format = self.percent_format
            
            row += 1
        
        # Key Metrics
        row += 1
        
        # Calculate Working Capital
        working_capital = [
            balance_sheet_data['total_current_assets'][i] - balance_sheet_data['total_current_liabilities'][i]
            for i in range(3)
        ]
        
        self.ws[f'C{row}'] = "Working Capital"
        self.ws[f'D{row}'] = self.format_currency(working_capital[0])
        self.ws[f'F{row}'] = self.calculate_percentage_change(working_capital[0], working_capital[1])
        self.ws[f'G{row}'] = self.format_currency(working_capital[1])
        self.ws[f'I{row}'] = self.calculate_percentage_change(working_capital[1], working_capital[2])
        self.ws[f'J{row}'] = self.format_currency(working_capital[2])
        self.ws[f'F{row}'].number_format = self.percent_format
        self.ws[f'I{row}'].number_format = self.percent_format
        row += 1
        
        # Net Worth (Shareholders' Equity)
        self.ws[f'C{row}'] = "Net Worth (OE)"
        self.ws[f'D{row}'] = self.format_currency(balance_sheet_data['total_equity'][0])
        self.ws[f'F{row}'] = self.calculate_percentage_change(balance_sheet_data['total_equity'][0], balance_sheet_data['total_equity'][1])
        self.ws[f'G{row}'] = self.format_currency(balance_sheet_data['total_equity'][1])
        self.ws[f'I{row}'] = self.calculate_percentage_change(balance_sheet_data['total_equity'][1], balance_sheet_data['total_equity'][2])
        self.ws[f'J{row}'] = self.format_currency(balance_sheet_data['total_equity'][2])
        self.ws[f'F{row}'].number_format = self.percent_format
        self.ws[f'I{row}'].number_format = self.percent_format
        row += 1
        
        # Current Ratio
        current_ratio = [
            balance_sheet_data['total_current_assets'][i] / balance_sheet_data['total_current_liabilities'][i]
            if balance_sheet_data['total_current_liabilities'][i] != 0 else 0
            for i in range(3)
        ]
        
        self.ws[f'C{row}'] = "Current Ratio"
        self.ws[f'F{row}'] = round(current_ratio[0], 2)
        self.ws[f'H{row}'] = round(current_ratio[1], 2)
        self.ws[f'J{row}'] = round(current_ratio[2], 2)
        row += 1
        
        # Quick Ratio
        quick_assets = [
            (balance_sheet_data['total_current_assets'][i] - balance_sheet_data['inventory'][i])
            for i in range(3)
        ]
        quick_ratio = [
            quick_assets[i] / balance_sheet_data['total_current_liabilities'][i]
            if balance_sheet_data['total_current_liabilities'][i] != 0 else 0
            for i in range(3)
        ]
        
        self.ws[f'C{row}'] = "Quick Ratio"
        self.ws[f'F{row}'] = round(quick_ratio[0], 2)
        self.ws[f'H{row}'] = round(quick_ratio[1], 2)
        self.ws[f'J{row}'] = round(quick_ratio[2], 2)
        row += 1
        
        # Income Statement Section
        row += 3
        self.ws[f'B{row}'] = "Income Statement:"
        self.ws[f'B{row}'].font = self.header_font
        row += 2
        
        income_items = [
            ("Revenue", income_stmt_data['revenue']),
            ("Cost of Revenue", income_stmt_data['cost_of_revenue']),
            ("Gross Profit", income_stmt_data['gross_profit']),
            ("Operating Expenses", income_stmt_data['operating_expenses']),
            ("Operating Income", income_stmt_data['operating_income']),
            ("EBIT", income_stmt_data['ebit']),
            ("Interest Expense", income_stmt_data['interest_expense']),
            ("Pretax Income", income_stmt_data['pretax_income']),
            ("Tax Expense", income_stmt_data['tax_expense']),
            ("Net Income", income_stmt_data['net_income']),
        ]
        
        for item_name, values in income_items:
            self.ws[f'D{row}'] = item_name
            self.ws[f'F{row}'] = self.format_currency(values[0])
            self.ws[f'H{row}'] = self.format_currency(values[1])
            self.ws[f'J{row}'] = self.format_currency(values[2])
            
            if item_name != "Revenue":  # Skip percentage for revenue
                self.ws[f'G{row}'] = self.calculate_percentage_change(values[0], values[1])
                self.ws[f'I{row}'] = self.calculate_percentage_change(values[1], values[2])
                self.ws[f'G{row}'].number_format = self.percent_format
                self.ws[f'I{row}'].number_format = self.percent_format
            
            row += 1
        
        # Cash Flow Section
        row += 3
        self.ws[f'B{row}'] = "Cash Flows:"
        self.ws[f'B{row}'].font = self.header_font
        row += 2
        
        # Operating Activities
        self.ws[f'C{row}'] = "Cash Flows-Operating Activities:"
        self.ws[f'C{row}'].font = self.subheader_font
        row += 1
        
        self.ws[f'D{row}'] = "Net Income"
        self.ws[f'F{row}'] = self.format_currency(income_stmt_data['net_income'][0])
        self.ws[f'H{row}'] = self.format_currency(income_stmt_data['net_income'][1])
        self.ws[f'J{row}'] = self.format_currency(income_stmt_data['net_income'][2])
        row += 1
        
        self.ws[f'D{row}'] = "Depreciation & Amortization"
        self.ws[f'F{row}'] = self.format_currency(cash_flow_data['depreciation'][0])
        self.ws[f'H{row}'] = self.format_currency(cash_flow_data['depreciation'][1])
        self.ws[f'J{row}'] = self.format_currency(cash_flow_data['depreciation'][2])
        row += 1
        
        self.ws[f'D{row}'] = "Net Cash Flow-Operating"
        self.ws[f'F{row}'] = self.format_currency(cash_flow_data['operating_cash_flow'][0])
        self.ws[f'G{row}'] = self.calculate_percentage_change(cash_flow_data['operating_cash_flow'][0], cash_flow_data['operating_cash_flow'][1])
        self.ws[f'H{row}'] = self.format_currency(cash_flow_data['operating_cash_flow'][1])
        self.ws[f'I{row}'] = self.calculate_percentage_change(cash_flow_data['operating_cash_flow'][1], cash_flow_data['operating_cash_flow'][2])
        self.ws[f'J{row}'] = self.format_currency(cash_flow_data['operating_cash_flow'][2])
        self.ws[f'G{row}'].number_format = self.percent_format
        self.ws[f'I{row}'].number_format = self.percent_format
        row += 2
        
        # Investing Activities
        self.ws[f'C{row}'] = "Cash Flows-Investing Activities:"
        self.ws[f'C{row}'].font = self.subheader_font
        row += 1
        
        self.ws[f'D{row}'] = "Capital Expenditures"
        self.ws[f'F{row}'] = self.format_currency(cash_flow_data['capex'][0])
        self.ws[f'H{row}'] = self.format_currency(cash_flow_data['capex'][1])
        self.ws[f'J{row}'] = self.format_currency(cash_flow_data['capex'][2])
        row += 1
        
        self.ws[f'D{row}'] = "Net Cash Flows-Investing"
        self.ws[f'F{row}'] = self.format_currency(cash_flow_data['investing_cash_flow'][0])
        self.ws[f'H{row}'] = self.format_currency(cash_flow_data['investing_cash_flow'][1])
        self.ws[f'J{row}'] = self.format_currency(cash_flow_data['investing_cash_flow'][2])
        row += 2
        
        # Financing Activities
        self.ws[f'C{row}'] = "Cash Flows-Financing Activities:"
        self.ws[f'C{row}'].font = self.subheader_font
        row += 1
        
        self.ws[f'D{row}'] = "Debt Repayment"
        self.ws[f'F{row}'] = self.format_currency(cash_flow_data['debt_repayment'][0])
        self.ws[f'H{row}'] = self.format_currency(cash_flow_data['debt_repayment'][1])
        self.ws[f'J{row}'] = self.format_currency(cash_flow_data['debt_repayment'][2])
        row += 1
        
        self.ws[f'D{row}'] = "Net Cash Flows-Financing"
        self.ws[f'F{row}'] = self.format_currency(cash_flow_data['financing_cash_flow'][0])
        self.ws[f'H{row}'] = self.format_currency(cash_flow_data['financing_cash_flow'][1])
        self.ws[f'J{row}'] = self.format_currency(cash_flow_data['financing_cash_flow'][2])
        row += 2
        
        # Net Cash Flow
        net_cash_flow = [
            cash_flow_data['operating_cash_flow'][i] + 
            cash_flow_data['investing_cash_flow'][i] + 
            cash_flow_data['financing_cash_flow'][i]
            for i in range(3)
        ]
        
        self.ws[f'D{row}'] = "Net Cash Flow"
        self.ws[f'F{row}'] = self.format_currency(net_cash_flow[0])
        self.ws[f'H{row}'] = self.format_currency(net_cash_flow[1])
        self.ws[f'J{row}'] = self.format_currency(net_cash_flow[2])
        
        # Apply column widths
        self.ws.column_dimensions['A'].width = 2
        self.ws.column_dimensions['B'].width = 4
        self.ws.column_dimensions['C'].width = 30
        self.ws.column_dimensions['D'].width = 35
        self.ws.column_dimensions['E'].width = 2
        self.ws.column_dimensions['F'].width = 18
        self.ws.column_dimensions['G'].width = 12
        self.ws.column_dimensions['H'].width = 18
        self.ws.column_dimensions['I'].width = 12
        self.ws.column_dimensions['J'].width = 18
        
    def save_report(self):
        """Save the Excel report"""
        self.wb.save(self.filename)
        logger.info(f"Report saved as {self.filename}")


class FinancialReportAutomation:
    """Main automation class that orchestrates the entire process"""
    
    def __init__(self, ticker="TSLA", output_file="tesla_financial_report.xlsx"):
        self.ticker = ticker
        self.output_file = output_file
        self.scraper = FinancialDataScraper(ticker)
        
    def generate_report(self):
        """Generate the complete financial report"""
        logger.info(f"Starting report generation for {self.ticker}...")
        
        # Fetch all financial data
        logger.info("Fetching balance sheet data...")
        balance_sheet_data = self.scraper.get_balance_sheet_data()
        
        logger.info("Fetching income statement data...")
        income_stmt_data = self.scraper.get_income_statement_data()
        
        logger.info("Fetching cash flow data...")
        cash_flow_data = self.scraper.get_cash_flow_data()
        
        # Check if data was successfully fetched
        if not all([balance_sheet_data, income_stmt_data, cash_flow_data]):
            logger.error("Failed to fetch all required financial data")
            return False
        
        # Generate Excel report
        logger.info("Generating Excel report...")
        report_generator = ExcelReportGenerator(self.output_file)
        report_generator.create_report(balance_sheet_data, income_stmt_data, cash_flow_data)
        report_generator.save_report()
        
        logger.info("Report generation completed successfully!")
        return True
    
    def schedule_daily_report(self, time_str="09:00"):
        """Schedule daily report generation"""
        schedule.every().day.at(time_str).do(self.generate_report)
        logger.info(f"Scheduled daily report generation at {time_str}")
        
        while True:
            schedule.run_pending()
            time.sleep(60)  # Check every minute


def main():
    """Main function to run the automation"""
    
    # Configuration
    TICKER = "TSLA"  # Tesla stock ticker
    OUTPUT_FILE = "tesla_financial_report.xlsx"
    
    # Create automation instance
    automation = FinancialReportAutomation(TICKER, OUTPUT_FILE)
    
    # Option 1: Generate report immediately
    automation.generate_report()
    
    # Option 2: Schedule daily reports (uncomment to use)
    # automation.schedule_daily_report("09:00")  # Run every day at 9 AM
    

if __name__ == "__main__":
    main()